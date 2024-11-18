#!/usr/bin/env python3

import requests
import json
import sys
import os
import threading
from queue import Queue
import urllib3
import urllib.parse
from html.parser import HTMLParser
import io
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import argparse
import logging

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')

contentList = []
queue = Queue()

lock = threading.Lock()
image_counter_lock = threading.Lock()

default_headers = {
    "Accept": "application/json",
    "User-Agent": "ConfluenceScraper"
}

image_counter = 0

def escape_cql_term(term):
    special_chars = ['\\', '"', "'", '=', '+', '-', '*', '?', ':', '{', '}', '[', ']', '(', ')', '^', '~', '!', '|', '&', '/']
    for char in special_chars:
        term = term.replace(char, f'\\{char}')
    return term

def getNumberOfPages(query, cURL, headers):
    q = "/rest/api/content/search"
    URL = cURL + q
    try:
        response = requests.get(URL, headers=headers, params=query, verify=False)
        response.raise_for_status()
        jsonResp = response.json()
        return int(jsonResp.get("totalSize", 0))
    except requests.RequestException as e:
        logging.error(f"Error during API request: {e}")
        return 0
    except (ValueError, json.JSONDecodeError):
        logging.error("Error parsing JSON response.")
        return 0

def fetchPageContent(page_id, cURL, headers):
    URL = f"{cURL}/rest/api/content/{page_id}?expand=body.storage"
    try:
        response = requests.get(URL, headers=headers, verify=False)
        response.raise_for_status()
        jsonResp = response.json()
        content = jsonResp['body']['storage']['value']
        return content
    except requests.RequestException as e:
        logging.error(f"Error fetching page content: {e}")
        return None
    except (ValueError, json.JSONDecodeError):
        logging.error("Error parsing JSON response.")
        return None

def fetchPageAttachments(page_id, cURL, headers):
    URL = f"{cURL}/rest/api/content/{page_id}/child/attachment"
    try:
        response = requests.get(URL, headers=headers, verify=False)
        response.raise_for_status()
        jsonResp = response.json()
        attachments = jsonResp.get('results', [])
        return attachments
    except requests.RequestException as e:
        logging.error(f"Error fetching attachments: {e}")
        return []
    except (ValueError, json.JSONDecodeError):
        logging.error("Error parsing JSON response.")
        return []

def downloadAttachment(attachment, cURL, headers):
    download_url = attachment['_links']['download']
    URL = cURL + download_url
    try:
        response = requests.get(URL, headers=headers, verify=False)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        logging.error(f"Error downloading attachment: {e}")
        return None

def extract_text_from_pdf(content):
    try:
        pdf_reader = PdfReader(io.BytesIO(content))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        logging.error(f"Error extracting text from PDF: {e}")
        return ""

def extract_text_from_docx(content):
    try:
        document = Document(io.BytesIO(content))
        text = "\n".join([para.text for para in document.paragraphs])
        return text
    except Exception as e:
        logging.error(f"Error extracting text from DOCX: {e}")
        return ""

def extract_text_from_txt(content):
    try:
        return content.decode('utf-8', errors='ignore')
    except Exception as e:
        logging.error(f"Error extracting text from TXT: {e}")
        return ""

def extract_text_from_xlsx(content):
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        logging.error(f"Error extracting text from XLSX: {e}")
        return ""

def extract_text_from_image(content):
    try:
        image = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        logging.error(f"Error extracting text from image: {e}")
        return ""

class HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.reset()
        self.fed = []

    def handle_data(self, d):
        self.fed.append(d)

    def get_data(self):
        return ''.join(self.fed)

def strip_tags(html):
    s = HTMLStripper()
    s.feed(html)
    return s.get_data()

def searchKeyWordsThread(queue, cURL, headers, limit, search_content, num_chars, search_binaries, search_images, max_images):
    global image_counter
    while not queue.empty():
        search_term = queue.get()
        search_term_decoded = urllib.parse.unquote(search_term)
        escaped_term = escape_cql_term(search_term_decoded)
        cql_query = f'text ~ "{escaped_term}"'
        totalSize = getNumberOfPages({'cql': cql_query}, cURL, headers)

        if totalSize:
            logging.info(f"{totalSize} results for search term: {search_term_decoded}")
            start_point = 0
            totalSize = min(totalSize, limit)

            while start_point < totalSize:
                logging.info(f"Processing {start_point} of {totalSize} results for term: {search_term_decoded}")
                searchQuery = {
                    'cql': cql_query,
                    'start': start_point,
                    'limit': 50
                }
                URL = cURL + "/rest/api/content/search"

                try:
                    response = requests.get(URL, headers=headers, params=searchQuery, verify=False)
                    response.raise_for_status()
                    jsonResp = response.json()

                    results = jsonResp.get('results', [])
                    if not results:
                        break

                    for result in results:
                        try:
                            page_id = result['id']
                            page_url = cURL + result['_links']['webui']
                            page_name = result['title']

                            if search_content:
                                content = fetchPageContent(page_id, cURL, headers)
                                if content:
                                    clean_content = strip_tags(content)

                                    index = clean_content.lower().find(search_term_decoded.lower())
                                    if index != -1:
                                        start_index = index + len(search_term_decoded)
                                        end_index = start_index + num_chars
                                        extracted_info = clean_content[start_index:end_index].strip()
                                        if not extracted_info:
                                            extracted_info = 'No extracted info'
                                    else:
                                        extracted_info = 'Search term not found in content'

                                    with lock:
                                        contentList.append((page_url, page_name, search_term_decoded, extracted_info))
                                else:
                                    logging.warning(f"Could not fetch content for page '{page_name}'")
                                    with lock:
                                        contentList.append((page_url, page_name, search_term_decoded, 'Content not fetched'))
                            else:
                                with lock:
                                    contentList.append((page_url, page_name))

                            if search_binaries or search_images:
                                attachments = fetchPageAttachments(page_id, cURL, headers)
                                for attachment in attachments:
                                    attachment_title = attachment.get('title', '')
                                    file_extension = os.path.splitext(attachment_title)[1].lower()
                                    is_image = file_extension in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif']
                                    is_binary = file_extension in ['.pdf', '.docx', '.txt', '.xlsx']
                                    if is_binary or (search_images and is_image):
                                        if is_image and max_images is not None:
                                            with image_counter_lock:
                                                if image_counter >= max_images:
                                                    continue
                                                image_counter += 1

                                        attachment_content = downloadAttachment(attachment, cURL, headers)
                                        if attachment_content:
                                            if search_content:
                                                if file_extension == '.pdf':
                                                    extracted_text = extract_text_from_pdf(attachment_content)
                                                elif file_extension == '.docx':
                                                    extracted_text = extract_text_from_docx(attachment_content)
                                                elif file_extension == '.txt':
                                                    extracted_text = extract_text_from_txt(attachment_content)
                                                elif file_extension == '.xlsx':
                                                    extracted_text = extract_text_from_xlsx(attachment_content)
                                                elif is_image:
                                                    extracted_text = extract_text_from_image(attachment_content)
                                                else:
                                                    extracted_text = ''

                                                if extracted_text:
                                                    index = extracted_text.lower().find(search_term_decoded.lower())
                                                    if index != -1:
                                                        start_index = index + len(search_term_decoded)
                                                        end_index = start_index + num_chars
                                                        extracted_info = extracted_text[start_index:end_index].strip()
                                                        if not extracted_info:
                                                            extracted_info = 'No extracted info'
                                                    else:
                                                        extracted_info = 'Search term not found in attachment'

                                                    with lock:
                                                        contentList.append((page_url, page_name, search_term_decoded, extracted_info, attachment_title))
                                                else:
                                                    logging.warning(f"Could not extract text from attachment '{attachment_title}'")
                                            else:
                                                with lock:
                                                    contentList.append((page_url, page_name, attachment_title))
                                        else:
                                            logging.warning(f"Attachment content is empty for '{attachment_title}'")
                        except KeyError as e:
                            logging.error(f"Missing key in result: {e}")

                    start_point += len(results)

                except requests.RequestException as e:
                    logging.error(f"Error fetching results: {e}")
                    break
                except (ValueError, json.JSONDecodeError):
                    logging.error("Error parsing JSON response.")
                    break
            queue.task_done()
        else:
            logging.info(f"No documents found for term: {search_term_decoded}")
            queue.task_done()

def saveContent(search_content, search_binaries):
    logging.info("Saving extracted information to file")
    save_path = './loot'
    if search_content or search_binaries:
        file_name = "confluence_extracted.txt"
    else:
        file_name = "confluence_content.txt"
    completeName = os.path.join(save_path, file_name)

    os.makedirs(save_path, exist_ok=True)

    try:
        with open(completeName, "w") as f:
            if len(contentList) == 0:
                logging.info("No content to write to file.")
            else:
                for item in contentList:
                    page_url, page_name = item[0], item[1]
                    f.write(f"URL: {page_url}\n")
                    f.write(f"Page Title: {page_name}\n")
                    if len(item) == 5:
                        search_term = item[2]
                        extracted_info = item[3]
                        attachment_title = item[4]
                        f.write(f"Attachment Title: {attachment_title}\n")
                        f.write(f"Search Term: {search_term}\n")
                        f.write(f"Extracted Info: {extracted_info}\n")
                    elif len(item) == 3:
                        attachment_title = item[2]
                        f.write(f"Attachment Title: {attachment_title}\n")
                    elif len(item) == 4:
                        search_term = item[2]
                        extracted_info = item[3]
                        f.write(f"Search Term: {search_term}\n")
                        f.write(f"Extracted Info: {extracted_info}\n")
                    f.write("-" * 50 + "\n")
                    logging.info(f"Writing info from {page_url} to file.")
            logging.info(f"Saved content to file: {completeName}")
    except Exception as e:
        logging.error(f"Error saving content to file: {e}")

def main():
    global image_counter

    parser = argparse.ArgumentParser(description='ConfluenceScraper searches and extracts information from Confluence pages and attachments.')

    parser.add_argument('-c', '--url', required=True, help='Target Confluence base URL, e.g., https://confluence.company.com')
    parser.add_argument('-p', '--accesstoken', help='Confluence API access token (optional if the instance allows anonymous access)')

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('-w', '--wordlist', help='Path to a wordlist file containing search terms')
    group.add_argument('-k', '--keyword', help='A single keyword to search for')

    parser.add_argument('-s', '--search-content', action='store_true', help='Enable extraction of content from pages and attachments')
    parser.add_argument('-b', '--search-binaries', action='store_true', help='Include binary attachments (PDF, DOCX, XLSX, TXT) in the search')
    parser.add_argument('-i', '--search-images', action='store_true', help='Include image attachments in the search and perform OCR')
    parser.add_argument('-t', '--threads', type=int, default=10, help='Number of threads to use (default: 10)')
    parser.add_argument('-l', '--limit', type=int, default=float('inf'), help='Limit the number of search results per keyword (default: no limit)')
    parser.add_argument('-n', '--num-chars', type=int, default=10, help='Number of characters to extract after the search term (default: 10)')
    parser.add_argument('-m', '--max-images', type=int, help='Maximum number of images to process (default: no limit)')

    args = parser.parse_args()

    cURL = args.url.rstrip('/')
    access_token = args.accesstoken
    wordlist_path = args.wordlist
    single_keyword = args.keyword
    thread_count = args.threads
    limit = args.limit
    search_content = args.search_content
    search_binaries = args.search_binaries
    search_images = args.search_images
    num_chars = args.num_chars
    max_images = args.max_images

    if access_token:
        default_headers['Authorization'] = f'Bearer {access_token}'

    logging.info("Starting ConfluenceScraper with the following parameters:")
    logging.info(f"URL: {cURL}")
    if access_token:
        logging.info("Using provided access token for authentication.")
    else:
        logging.info("No access token provided; attempting anonymous access.")
    if wordlist_path:
        logging.info(f"Using wordlist file: {wordlist_path}")
    elif single_keyword:
        logging.info(f"Using single keyword: {single_keyword}")
    logging.info(f"Threads: {thread_count}")
    logging.info(f"Limit per keyword: {limit}")
    logging.info(f"Search Content: {'Enabled' if search_content else 'Disabled'}")
    logging.info(f"Search Binaries: {'Enabled' if search_binaries else 'Disabled'}")
    logging.info(f"Search Images: {'Enabled' if search_images else 'Disabled'}")
    if search_images and max_images is not None:
        logging.info(f"Maximum Images to Process: {max_images}")
    logging.info(f"Number of Characters to Extract: {num_chars}")

    if wordlist_path:
        try:
            with open(wordlist_path, "r") as f:
                for line in f:
                    keyword = line.strip()
                    if keyword:
                        queue.put(keyword)
        except Exception as e:
            logging.error(f"Error opening wordlist file: {e}")
            sys.exit(2)
    elif single_keyword:
        queue.put(single_keyword)

    threads = []
    for _ in range(thread_count):
        thread = threading.Thread(target=searchKeyWordsThread, args=(queue, cURL, default_headers, limit, search_content, num_chars, search_binaries, search_images, max_images))
        thread.daemon = True
        threads.append(thread)
        thread.start()

    queue.join()

    saveContent(search_content, search_binaries)

if __name__ == "__main__":
    main()